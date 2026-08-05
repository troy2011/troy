import json
import sys
from pathlib import Path

from openpyxl import load_workbook


SUITS = {"カップ": "Cup", "ワンド": "Wand", "ソード": "Sword", "ペンタクル": "Pentacle"}
RANKS = {"A": 1, "P": 11, "N": 12, "Q": 13, "K": 14}
ATTRIBUTES = {"無": "neutral", "光": "light", "闇": "dark"}


def value(row, index):
    item = row[index]
    return "" if item is None else str(item)


def rank_value(raw):
    return int(RANKS.get(str(raw), raw))


def main():
    if len(sys.argv) != 4:
        raise SystemExit("usage: generate-tarot-kingdom-effects-v3.py workbook current-v2 output")
    workbook_path, current_path, output_path = map(Path, sys.argv[1:])
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    current = json.loads(current_path.read_text(encoding="utf-8"))
    current_attributes = {int(entry["number"]): entry.get("attribute", "neutral") for entry in current["guardian"]}

    r_sheet = workbook["数字共通参照"]
    r_by_rank = {}
    for row in r_sheet.iter_rows(min_row=5, max_row=18, min_col=1, max_col=8, values_only=True):
        rank = rank_value(row[0])
        r_by_rank[rank] = {
            "source": value(row, 2),
            "basis": value(row, 3),
            "formula": value(row, 5),
            "reset": value(row, 6),
        }

    minor_sheet = workbook["小アルカナ共鳴"]
    minor = []
    for row in minor_sheet.iter_rows(min_row=5, max_row=60, min_col=1, max_col=13, values_only=True):
        rank = rank_value(row[2])
        effect_id = value(row, 0)
        minor.append({
            "id": effect_id,
            "suit": SUITS[value(row, 1)],
            "rank": rank,
            "name": value(row, 7) or value(row, 3),
            "effect": value(row, 9) or value(row, 5),
            "range": value(row, 10),
            "r": r_by_rank[rank],
            "condition": {"kind": "resonance-v3"},
            "steps": [{"kind": "r-effect", "effectId": effect_id}],
        })

    guardian_sheet = workbook["守護アルカナ"]
    guardian = []
    for row in guardian_sheet.iter_rows(min_row=5, max_row=26, min_col=1, max_col=12, values_only=True):
        number = int(row[0])
        guardian.append({
            "number": number,
            "attribute": ATTRIBUTES.get(value(row, 6), current_attributes.get(number, "neutral")),
            "passiveId": f"guardian-v3-{number}",
            "passiveName": value(row, 7) or value(row, 3),
            "passive": value(row, 8) or value(row, 4),
        })

    guardian_attributes = {entry["number"]: entry["attribute"] for entry in guardian}
    major_sheet = workbook["大アルカナ固有"]
    major = []
    for row in major_sheet.iter_rows(min_row=5, max_row=26, min_col=1, max_col=11, values_only=True):
        number = int(row[0])
        major.append({
            "number": number,
            "attribute": guardian_attributes.get(number, "neutral"),
            "skillName": value(row, 6) or value(row, 2),
            "effect": value(row, 7) or value(row, 3),
            "effectId": f"major-v3-{number}",
        })

    if len(minor) != 56 or len(guardian) != 22 or len(major) != 22:
        raise RuntimeError(f"unexpected catalog counts: {len(minor)}, {len(guardian)}, {len(major)}")
    output_path.write_text(
        json.dumps({"version": 3, "minor": minor, "guardian": guardian, "major": major}, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


if __name__ == "__main__":
    main()
